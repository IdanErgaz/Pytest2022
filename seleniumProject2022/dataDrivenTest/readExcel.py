import openpyxl
#need to create the excel file from advance !!!!

path = 'C:\ExcelsFiles4Auto\Demo.xlsx'
workbook= openpyxl.load_workbook(path)
sheet= workbook.active  #workbook.get_sheet_by_name('Sheet1')

rows = sheet.max_row
cols= sheet.max_column
print('rows:{}'.format(rows))
print('cols:{}'.format(cols))

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r, column= c).value, end= "    ")
    print()